import os
import tempfile
import json
from decimal import Decimal
from io import BytesIO
from itertools import chain
from pathlib import Path

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.core.management import call_command
from django.db import connections
from django.db.models import Count
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views import generic
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .forms import (
    DatabaseRestoreForm,
    IndustrialDesignForm,
    RecordFilterForm,
    SiteSettingsForm,
    TrademarkForm,
)
from .models import IndustrialDesign, SiteSettings, Trademark
from .services.database_backup import BackupOperationError, DatabaseBackupService
from .utils import register_arabic_font, shape_arabic, with_total_fees


def healthcheck_view(request):
    return HttpResponse("ok", content_type="text/plain; charset=utf-8")


def run_post_restore_steps():
    connections.close_all()
    call_command("migrate", "--noinput", verbosity=0)


def get_backup_service():
    return DatabaseBackupService(
        database_path=Path(settings.DATABASES["default"]["NAME"]),
        backup_dir=Path(settings.BACKUP_DIR),
        media_root=Path(settings.MEDIA_ROOT),
        close_connections=connections.close_all,
        after_restore=run_post_restore_steps,
        filename_prefix=getattr(settings, "BACKUP_FILENAME_PREFIX", "brandregistry-backup"),
    )


def save_uploaded_backup_to_temp_file(uploaded_file) -> Path:
    suffix = Path(uploaded_file.name).suffix or ".sqlite3"
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)

    try:
        for chunk in uploaded_file.chunks():
            temp_file.write(chunk)
        temp_file.flush()
    finally:
        temp_file.close()

    return Path(temp_file.name)


def parse_json_request_body(request):
    content_type = request.headers.get("Content-Type", "")
    if "application/json" not in content_type:
        return {}

    if not request.body:
        return {}

    try:
        return json.loads(request.body.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise BackupOperationError("تعذر قراءة بيانات الطلب المرسلة من نسخة سطح المكتب.") from exc


def create_backup_response(target_directory: str | None = None):
    backup_record = get_backup_service().create_backup(target_directory)
    return {
        "ok": True,
        "message": f"تم إنشاء نسخة احتياطية باسم {backup_record.name}.",
        "backupName": backup_record.name,
        "backupPath": str(backup_record.path),
    }


def apply_record_filters(queryset, cleaned_data):
    queryset = with_total_fees(queryset)

    search_name = cleaned_data.get("search_name")
    search_number = cleaned_data.get("search_number")
    status = cleaned_data.get("status")
    filing_date_from = cleaned_data.get("filing_date_from")
    filing_date_to = cleaned_data.get("filing_date_to")
    expiring_only = cleaned_data.get("expiring_only")
    fees_min = cleaned_data.get("fees_min")
    fees_max = cleaned_data.get("fees_max")

    if search_name:
        queryset = queryset.filter(name__icontains=search_name)
    if search_number:
        queryset = queryset.filter(number__icontains=search_number)
    if status:
        queryset = queryset.filter(status=status)
    if filing_date_from:
        queryset = queryset.filter(filing_date__gte=filing_date_from)
    if filing_date_to:
        queryset = queryset.filter(filing_date__lte=filing_date_to)
    if expiring_only:
        matching_ids = [item.pk for item in queryset if item.is_approaching_protection_expiry]
        queryset = queryset.filter(pk__in=matching_ids)
    if fees_min is not None:
        queryset = queryset.filter(total_fees_db__gte=fees_min)
    if fees_max is not None:
        queryset = queryset.filter(total_fees_db__lte=fees_max)

    return queryset


def build_status_counts(model):
    counts = {key: 0 for key, _ in model.STATUS_CHOICES}
    counts.update(model.objects.values("status").annotate(total=Count("id")).values_list("status", "total"))
    return counts


def collect_protection_groups(records):
    expiring = []
    expired = []
    renewal_due = []

    for item in records:
        if item.needs_renewal:
            renewal_due.append(item)
        if item.is_expired:
            expired.append(item)
        elif item.is_approaching_protection_expiry:
            expiring.append(item)

    return {
        "expiring": expiring,
        "expired": expired,
        "renewal_due": renewal_due,
    }


class DashboardView(generic.TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        trademark_counts = build_status_counts(Trademark)
        design_counts = build_status_counts(IndustrialDesign)
        all_trademarks = list(Trademark.objects.all())
        all_designs = list(IndustrialDesign.objects.all())
        trademark_records = all_trademarks[:6]
        design_records = all_designs[:6]
        trademark_groups = collect_protection_groups(all_trademarks)
        design_groups = collect_protection_groups(all_designs)

        dashboard_notifications = [
            {
                "title": "علامات بحاجة إلى تجديد",
                "count": len(trademark_groups["renewal_due"]),
                "tone": "danger",
                "description": "انتهت مدة الحماية وتحتاج إلى تسجيل التجديد.",
            },
            {
                "title": "نماذج بحاجة إلى تجديد",
                "count": len(design_groups["renewal_due"]),
                "tone": "danger",
                "description": "انتهت مدة الحماية وتحتاج إلى تسجيل التجديد.",
            },
            {
                "title": "علامات منتهية الحماية",
                "count": len(trademark_groups["expired"]),
                "tone": "danger",
                "description": "سجلات انتهت مدة الحماية الخاصة بها بالفعل.",
            },
            {
                "title": "نماذج منتهية الحماية",
                "count": len(design_groups["expired"]),
                "tone": "danger",
                "description": "سجلات انتهت مدة الحماية الخاصة بها بالفعل.",
            },
            {
                "title": "علامات على وشك الانتهاء",
                "count": len(trademark_groups["expiring"]),
                "tone": "warning",
                "description": "سجلات اقترب انتهاء حمايتها خلال 30 يومًا.",
            },
            {
                "title": "نماذج على وشك الانتهاء",
                "count": len(design_groups["expiring"]),
                "tone": "warning",
                "description": "سجلات اقترب انتهاء حمايتها خلال 30 يومًا.",
            },
        ]
        current_alert_keys = {
            ("trademark", item.pk)
            for item in trademark_groups["expiring"] + trademark_groups["expired"]
        } | {
            ("design", item.pk)
            for item in design_groups["expiring"] + design_groups["expired"]
        }

        recent_records = sorted(
            chain(
                [{"type": "علامة تجارية", "object": item} for item in trademark_records],
                [{"type": "نموذج صناعي", "object": item} for item in design_records],
            ),
            key=lambda item: item["object"].created_at,
            reverse=True,
        )[:8]

        context.update(
            {
                "trademark_counts": trademark_counts,
                "design_counts": design_counts,
                "total_trademarks": Trademark.objects.count(),
                "total_designs": IndustrialDesign.objects.count(),
                "total_registered_designs": design_counts[IndustrialDesign.STATUS_REGISTERED],
                "expiring_trademarks": trademark_groups["expiring"],
                "expired_trademarks": trademark_groups["expired"],
                "renewal_due_trademarks": trademark_groups["renewal_due"],
                "expiring_designs": design_groups["expiring"],
                "expired_designs": design_groups["expired"],
                "renewal_due_designs": design_groups["renewal_due"],
                "current_alerts": len(current_alert_keys),
                "dashboard_notifications": dashboard_notifications,
                "recent_records": recent_records,
                "total_fees_trademarks": sum(
                    (item.total_fees for item in Trademark.objects.all()),
                    Decimal("0.00"),
                ),
                "total_fees_designs": sum(
                    (item.total_fees for item in IndustrialDesign.objects.all()),
                    Decimal("0.00"),
                ),
                "total_renewal_fees_trademarks": sum(
                    (item.renewal_fee for item in Trademark.objects.all()),
                    Decimal("0.00"),
                ),
                "total_renewal_fees_designs": sum(
                    (item.renewal_fee for item in IndustrialDesign.objects.all()),
                    Decimal("0.00"),
                ),
            }
        )
        return context


class BaseRecordListView(generic.ListView):
    paginate_by = 20
    filter_form_class = RecordFilterForm

    def get_filter_form(self):
        return self.filter_form_class(self.request.GET or None)

    def get_queryset(self):
        queryset = self.model.objects.all()
        self.filter_form = self.get_filter_form()
        if self.filter_form.is_valid():
            queryset = apply_record_filters(queryset, self.filter_form.cleaned_data)
        else:
            queryset = with_total_fees(queryset)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_form"] = self.filter_form
        context["status_counts"] = build_status_counts(self.model)
        context["active_filters"] = {
            key: value
            for key, value in self.request.GET.items()
            if value not in ("", None)
        }
        return context


class TrademarkListView(BaseRecordListView):
    model = Trademark
    template_name = "trademarks.html"
    context_object_name = "trademarks"


class TrademarkDetailView(generic.DetailView):
    model = Trademark
    template_name = "trademark_detail.html"
    context_object_name = "trademark"


class TrademarkCreateView(generic.CreateView):
    model = Trademark
    form_class = TrademarkForm
    template_name = "trademark_form.html"
    success_url = reverse_lazy("registry:trademarks:list")

    def form_valid(self, form):
        messages.success(self.request, "تم حفظ العلامة التجارية بنجاح.")
        return super().form_valid(form)


class TrademarkUpdateView(generic.UpdateView):
    model = Trademark
    form_class = TrademarkForm
    template_name = "trademark_form.html"
    success_url = reverse_lazy("registry:trademarks:list")

    def form_valid(self, form):
        messages.success(self.request, "تم تحديث بيانات العلامة التجارية بنجاح.")
        return super().form_valid(form)


class TrademarkDeleteView(generic.DeleteView):
    model = Trademark
    template_name = "trademark_confirm_delete.html"
    success_url = reverse_lazy("registry:trademarks:list")

    def form_valid(self, form):
        messages.success(self.request, "تم حذف العلامة التجارية بنجاح.")
        return super().form_valid(form)


class DesignListView(BaseRecordListView):
    model = IndustrialDesign
    template_name = "designs.html"
    context_object_name = "designs"


class DesignDetailView(generic.DetailView):
    model = IndustrialDesign
    template_name = "design_detail.html"
    context_object_name = "design"


class DesignCreateView(generic.CreateView):
    model = IndustrialDesign
    form_class = IndustrialDesignForm
    template_name = "design_form.html"
    success_url = reverse_lazy("registry:designs:list")

    def form_valid(self, form):
        messages.success(self.request, "تم حفظ النموذج الصناعي بنجاح.")
        return super().form_valid(form)


class DesignUpdateView(generic.UpdateView):
    model = IndustrialDesign
    form_class = IndustrialDesignForm
    template_name = "design_form.html"
    success_url = reverse_lazy("registry:designs:list")

    def form_valid(self, form):
        messages.success(self.request, "تم تحديث بيانات النموذج الصناعي بنجاح.")
        return super().form_valid(form)


class DesignDeleteView(generic.DeleteView):
    model = IndustrialDesign
    template_name = "design_confirm_delete.html"
    success_url = reverse_lazy("registry:designs:list")

    def form_valid(self, form):
        messages.success(self.request, "تم حذف النموذج الصناعي بنجاح.")
        return super().form_valid(form)


class ReportsView(generic.TemplateView):
    template_name = "reports.html"

    report_config = {
        "trademark": {
            "model": Trademark,
            "title": "العلامات التجارية",
        },
        "design": {
            "model": IndustrialDesign,
            "title": "النماذج الصناعية",
        },
    }

    def dispatch(self, request, *args, **kwargs):
        self.record_type = request.GET.get("record_type", "trademark")
        self.config = self.report_config.get(self.record_type, self.report_config["trademark"])
        self.filter_form = RecordFilterForm(request.GET or None)
        self.queryset = self.config["model"].objects.all()
        if self.filter_form.is_valid():
            self.queryset = apply_record_filters(self.queryset, self.filter_form.cleaned_data)
        else:
            self.queryset = with_total_fees(self.queryset)

        export = request.GET.get("export")
        if export in {"pdf", "excel"}:
            return self.export_report(export)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        trademark_groups = collect_protection_groups(list(Trademark.objects.all()))
        design_groups = collect_protection_groups(list(IndustrialDesign.objects.all()))
        context.update(
            {
                "record_type": self.record_type,
                "record_title": self.config["title"],
                "filter_form": self.filter_form,
                "records": self.queryset[:100],
                "records_count": self.queryset.count(),
                "trademark_counts": build_status_counts(Trademark),
                "design_counts": build_status_counts(IndustrialDesign),
                "expiring_trademarks": trademark_groups["expiring"],
                "expired_trademarks": trademark_groups["expired"],
                "expiring_designs": design_groups["expiring"],
                "expired_designs": design_groups["expired"],
            }
        )
        return context

    def export_report(self, export_format):
        title = f"تقرير {self.config['title']}"
        records = list(self.queryset)
        rows = []
        for record in records:
            rows.append(
                [
                    record.name,
                    record.number,
                    record.get_status_display(),
                    record.filing_date.strftime("%Y-%m-%d") if record.filing_date else "",
                    record.decision_date.strftime("%Y-%m-%d") if record.decision_date else "",
                    record.publication_deadline.strftime("%Y-%m-%d") if record.publication_deadline else "",
                    record.protection_expiry.strftime("%Y-%m-%d") if record.protection_expiry else "",
                    record.protection_status_label,
                    str(record.renewal_count),
                    record.last_renewal_date.strftime("%Y-%m-%d") if record.last_renewal_date else "",
                    record.renewal_status,
                    f"{record.renewal_fee:.2f}",
                    f"{record.total_fees:.2f}",
                ]
            )

        headers = [
            "الاسم",
            "الرقم",
            "القرار",
            "تاريخ الإيداع",
            "تاريخ القرار",
            "آخر يوم في النشر",
            "انتهاء الحماية",
            "حالة الحماية",
            "عدد التجديدات",
            "تاريخ آخر تجديد",
            "حالة التجديد",
            "رسوم التجديد",
            "إجمالي الرسوم",
        ]

        if export_format == "excel":
            return self.export_excel(title, headers, rows)
        return self.export_pdf(title, headers, rows, records)

    def export_excel(self, title, headers, rows):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = title[:31]
        sheet.sheet_view.rightToLeft = True

        header_fill = PatternFill("solid", fgColor="0F4C5C")
        header_font = Font(color="FFFFFF", bold=True)
        right_align = Alignment(horizontal="right", vertical="center")

        sheet.append(headers)
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = right_align

        for row in rows:
            sheet.append(row)

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 4, 28)
            for cell in column_cells:
                cell.alignment = right_align

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{self.record_type}-report.xlsx"'
        return response

    def build_pdf_image_cell(self, record, text_style):
        image_field = getattr(record, "image", None)
        image_path = getattr(image_field, "path", "") if image_field else ""
        if not image_path:
            return Paragraph(shape_arabic("—"), text_style)

        try:
            image_reader = ImageReader(image_path)
            width, height = image_reader.getSize()
            if not width or not height:
                raise ValueError("Invalid image dimensions.")

            max_width = 34
            max_height = 34
            scale = min(max_width / width, max_height / height)
            thumbnail = RLImage(
                image_path,
                width=max(1, width * scale),
                height=max(1, height * scale),
            )
            thumbnail.hAlign = "CENTER"
            return thumbnail
        except Exception:
            return Paragraph(shape_arabic("—"), text_style)

    def export_pdf(self, title, headers, rows, records):
        font_name, bold_name = register_arabic_font()
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=18,
            leftMargin=18,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ArabicTitle",
            parent=styles["Title"],
            fontName=bold_name,
            fontSize=18,
            alignment=1,
        )
        cell_style = ParagraphStyle(
            "ArabicCell",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=9,
            leading=12,
            alignment=2,
        )
        header_style = ParagraphStyle(
            "ArabicHeader",
            parent=cell_style,
            fontName=bold_name,
            textColor=colors.white,
        )

        pdf_headers = list(reversed([*headers, "الصورة المرفقة"]))

        data = [[Paragraph(shape_arabic(header), header_style) for header in pdf_headers]]
        for record, row in zip(records, rows):
            row_with_image = [*row, self.build_pdf_image_cell(record, cell_style)]
            pdf_row = list(reversed(row_with_image))
            data.append(
                [
                    Paragraph(shape_arabic(cell), cell_style) if isinstance(cell, str) else cell
                    for cell in pdf_row
                ]
            )

        table = Table(data, repeatRows=1, colWidths=[60] + [None] * (len(pdf_headers) - 1))
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F4C5C")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D7DE")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FA")]),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ]
            )
        )

        elements = [
            Paragraph(shape_arabic(title), title_style),
            Spacer(1, 12),
            table,
        ]
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{self.record_type}-report.pdf"'
        return response


class SettingsView(generic.UpdateView):
    model = SiteSettings
    form_class = SiteSettingsForm
    template_name = "settings.html"
    success_url = reverse_lazy("registry:settings")

    def get_object(self, queryset=None):
        return SiteSettings.get_solo()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        backup_service = get_backup_service()
        context["restore_form"] = kwargs.get("restore_form") or DatabaseRestoreForm()
        context["backup_records"] = backup_service.list_backups()
        context["backup_directory"] = settings.BACKUP_DIR
        return context

    def form_valid(self, form):
        messages.success(self.request, "تم تحديث إعدادات النظام بنجاح.")
        return super().form_valid(form)


class DatabaseBackupCreateView(generic.View):
    def post(self, request, *args, **kwargs):
        try:
            backup_record = get_backup_service().create_backup()
        except BackupOperationError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(
                request,
                f"تم إنشاء نسخة احتياطية جديدة باسم {backup_record.name} داخل المجلد {settings.BACKUP_DIR}.",
            )

        return redirect("registry:settings")


@method_decorator(csrf_exempt, name="dispatch")
class DesktopBackupCreateView(generic.View):
    def post(self, request, *args, **kwargs):
        if not getattr(settings, "DESKTOP_LOCAL_MODE", False):
            return JsonResponse({"ok": False, "message": "هذا المسار متاح فقط داخل نسخة سطح المكتب."}, status=403)

        if request.headers.get("X-Desktop-App") != "1":
            return JsonResponse({"ok": False, "message": "الطلب غير مصرح به."}, status=403)

        try:
            payload = parse_json_request_body(request)
            response_payload = create_backup_response(payload.get("targetDirectory"))
        except BackupOperationError as exc:
            return JsonResponse({"ok": False, "message": str(exc)}, status=500)

        return JsonResponse(response_payload)


class DatabaseBackupRestoreView(generic.View):
    def post(self, request, *args, **kwargs):
        form = DatabaseRestoreForm(request.POST, request.FILES)

        if not form.is_valid():
            for field_errors in form.errors.values():
                for error in field_errors:
                    messages.error(request, error)
            return redirect("registry:settings")

        temp_backup_path = None
        try:
            temp_backup_path = save_uploaded_backup_to_temp_file(form.cleaned_data["backup_file"])
            get_backup_service().restore_from_backup(temp_backup_path)
        except BackupOperationError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(request, "تمت استعادة قاعدة البيانات من النسخة الاحتياطية بنجاح.")
        finally:
            if temp_backup_path and temp_backup_path.exists():
                os.unlink(temp_backup_path)

        return redirect("registry:settings")


class DatabaseBackupDownloadView(generic.View):
    def get(self, request, backup_name, *args, **kwargs):
        try:
            backup_path = get_backup_service().resolve_backup_path(backup_name)
        except BackupOperationError as exc:
            raise Http404(str(exc)) from exc

        return FileResponse(backup_path.open("rb"), as_attachment=True, filename=backup_path.name)


@method_decorator(csrf_exempt, name="dispatch")
class DesktopBackupBeforeExitView(generic.View):
    def post(self, request, *args, **kwargs):
        if not getattr(settings, "DESKTOP_LOCAL_MODE", False):
            return JsonResponse({"ok": False, "message": "هذا المسار متاح فقط داخل نسخة سطح المكتب."}, status=403)

        if request.headers.get("X-Desktop-App") != "1":
            return JsonResponse({"ok": False, "message": "الطلب غير مصرح به."}, status=403)

        try:
            payload = parse_json_request_body(request)
            response_payload = create_backup_response(payload.get("targetDirectory"))
        except BackupOperationError as exc:
            return JsonResponse({"ok": False, "message": str(exc)}, status=500)

        return JsonResponse(response_payload)


class NotFoundView(generic.TemplateView):
    template_name = "404.html"
